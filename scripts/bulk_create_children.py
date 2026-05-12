"""
Bulk Create Client Child Work Items from Parent ADO Work Items.

Self-contained one-off script that queries ADO for parent work items and creates
the corresponding child client work items — WITHOUT touching HappyFox.

Replicates the same child-creation logic as the live integration
(ado_service.create_child_work_item) but has zero dependency on the integration
config, settings, or any HappyFox configuration.

The Client Selection for Portal Syncing field on each parent determines which
clients the children are created for. Parents with an empty portal field are
skipped. Existing children are detected via ADO parent-child relations — if a
parent already has a child of the correct type for that client, it is skipped.

Outputs an Excel spreadsheet mapping Parent ID/URL → Child ID/URL.

Usage
-----
  # Dry run — preview which children would be created
  python scripts/bulk_create_children.py \\
    --project "MyProject" \\
    --types "User Story,Bug" \\
    --states "Ready for Development,In Progress" \\
    --dry-run

  # Create children and generate report
  python scripts/bulk_create_children.py \\
    --project "MyProject" \\
    --types "User Story" \\
    --states "Ready for Development" \\
    --output bulk_children_report.xlsx

  # Run from a saved ADO query (no --types or --states needed)
  python scripts/bulk_create_children.py \\
    --project "MyProject" \\
    --query-id "a1b2c3d4-e5f6-7890-abcd-ef1234567890" \\
    --dry-run

  # Override organization
  python scripts/bulk_create_children.py \\
    --org BrandtInfoServices \\
    --project "MyProject" \\
    --types "User Story" \\
    --states "New"

Authentication
--------------
Uses DefaultAzureCredential (Azure CLI locally, Managed Identity in Azure).
Requires: `az login` before running locally.

Dependencies
------------
  pip install httpx openpyxl azure-identity
"""

from __future__ import annotations

import argparse
import logging
import os
import time
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("bulk_create_children")

# ADO API version
ADO_API_VERSION = "7.1"

# Azure AD resource ID for ADO
ADO_RESOURCE_ID = "499b84ac-1321-427f-aa17-267ca6975798"

# Refresh token 5 minutes before actual expiry
TOKEN_REFRESH_BUFFER_SECONDS = 300

# Default organization (can be overridden with --org)
DEFAULT_ORG = os.environ.get("ADO_ORGANIZATION", "BrandtInfoServices")


# ---------------------------------------------------------------------------
# ADO field reference name constants (inlined — no dependency on config.py)
# ---------------------------------------------------------------------------

# Standard fields
TITLE = "System.Title"
DESCRIPTION = "System.Description"
STATE = "System.State"
WORK_ITEM_TYPE = "System.WorkItemType"
TEAM_PROJECT = "System.TeamProject"

# Standard extended
PRIORITY = "Microsoft.VSTS.Common.Priority"
SEVERITY = "Microsoft.VSTS.Common.Severity"
ACCEPTANCE_CRITERIA = "Microsoft.VSTS.Common.AcceptanceCriteria"
REPRO_STEPS = "Microsoft.VSTS.TCM.ReproSteps"

# Custom fields — defaults match the current ADO process template.
# Override via env vars if your field references differ.
CLIENT_REQUESTED = os.environ.get("ADO_FIELD_CLIENT_REQUESTED", "Custom.ClientRequested")
CLIENT_SELECTION_PORTAL = os.environ.get("ADO_FIELD_CLIENT_SELECTION_PORTAL", "Custom.ClientSelectionforPortalVisibility")
REQUEST_CATEGORY = os.environ.get("ADO_FIELD_REQUEST_CATEGORY", "Custom.RequestCategory")
TEST_SCENARIOS = os.environ.get("ADO_FIELD_TEST_SCENARIOS", "Custom.TestScenarios")
RELEASE_NOTES = os.environ.get("ADO_FIELD_RELEASE_NOTES", "Custom.ReleaseNotes")
SCRUM_TEAM = os.environ.get("ADO_FIELD_SCRUM_TEAM", "Custom.ScrumTeam")


# ---------------------------------------------------------------------------
# Work item type mappings (inlined — same as ado_models.py)
# ---------------------------------------------------------------------------

PARENT_TO_CHILD: dict[str, str] = {
    "User Story": "Client Story Work Item",
    "Issue": "Client Story Work Item",
    "Epic": "Client Epic Work Item",
    "Bug": "Client Bug Work Item",
    "Initiative": "Client Initiative Work Item",
    "Feature": "Client Feature Work Item",
}

ALL_PARENT_TYPES: set[str] = set(PARENT_TO_CHILD.keys())

# Parent ADO state → child ADO state. Only types with non-1:1 mappings are
# listed. User Story and Bug have identical parent/child states so they're
# omitted — the parent state is copied as-is.
STATE_TO_CHILD_STATE: dict[str, dict[str, str]] = {
    "Epic": {
        "New": "New",
        "Backlog": "New",
        "Open": "In Progress",
        "Resolved": "Completed",
        "In Progress": "In Progress",
        "Completed": "Completed",
        "Removed": "Removed",
    },
    "Initiative": {
        "New": "New",
        "Backlog": "Backlog",
        "In Progress": "In Progress",
        "Completed": "Completed",
        "Removed": "Removed",
    },
    "Feature": {
        "New": "New",
        "Backlog": "Backlog",
        "In Progress": "In Progress",
        "Completed": "Completed",
        "Removed": "Removed",
    },
}


# ---------------------------------------------------------------------------
# Per-parent-type field sync config (same as ado_service.py)
# ---------------------------------------------------------------------------

# (content_fields, metadata_fields) — content fields are copied as-is here
# (the live integration filters them through the [TAG] content parser, but
# for a one-off bulk create the full content is fine).

_SYNC_FIELDS: dict[str, tuple[list[str], list[str]]] = {
    "User Story": (
        [DESCRIPTION, ACCEPTANCE_CRITERIA, TEST_SCENARIOS, RELEASE_NOTES],
        [REQUEST_CATEGORY, PRIORITY, SCRUM_TEAM],
    ),
    "Issue": (
        [DESCRIPTION, ACCEPTANCE_CRITERIA, TEST_SCENARIOS, RELEASE_NOTES],
        [REQUEST_CATEGORY, PRIORITY, SCRUM_TEAM],
    ),
    "Bug": (
        [DESCRIPTION, REPRO_STEPS, TEST_SCENARIOS, RELEASE_NOTES],
        [REQUEST_CATEGORY, PRIORITY, SCRUM_TEAM, SEVERITY],
    ),
    "Feature": (
        [DESCRIPTION, ACCEPTANCE_CRITERIA],
        [REQUEST_CATEGORY, SCRUM_TEAM, PRIORITY],
    ),
}

# Default for types not explicitly listed (Epic, Initiative) — matches User Story.
_DEFAULT_SYNC = _SYNC_FIELDS["User Story"]


# ---------------------------------------------------------------------------
# Lightweight ADO HTTP client (no dependency on ado_service.py or config.py)
# ---------------------------------------------------------------------------

class AdoClient:
    """Minimal ADO REST API client using DefaultAzureCredential."""

    def __init__(self, org: str) -> None:
        self._org = org
        self._base_url = f"https://dev.azure.com/{org}/_apis"
        self._credential = None
        self._cached_token: str | None = None
        self._token_expires_at: float = 0.0
        self._client = httpx.Client(
            headers={"Content-Type": "application/json-patch+json"},
            timeout=30.0,
            event_hooks={"request": [self._inject_auth]},
        )

    def _inject_auth(self, request: httpx.Request) -> None:
        token = self._get_token()
        request.headers["Authorization"] = f"Bearer {token}"

    def _get_token(self) -> str:
        now = time.time()
        if self._cached_token and now < self._token_expires_at:
            return self._cached_token

        if self._credential is None:
            from azure.identity import DefaultAzureCredential
            self._credential = DefaultAzureCredential()

        token = self._credential.get_token(f"{ADO_RESOURCE_ID}/.default")
        self._cached_token = token.token
        self._token_expires_at = token.expires_on - TOKEN_REFRESH_BUFFER_SECONDS
        logger.info("Acquired ADO bearer token (expires in %ds)", int(token.expires_on - now))
        return self._cached_token

    def _project_url(self, project: str) -> str:
        return f"https://dev.azure.com/{self._org}/{project}/_apis"

    def close(self) -> None:
        self._client.close()

    # -- API methods --

    def get_work_item(self, work_item_id: int, expand: str = "all") -> dict[str, Any]:
        url = f"{self._base_url}/wit/workitems/{work_item_id}"
        params = {"api-version": ADO_API_VERSION, "$expand": expand}
        resp = self._client.get(url, params=params)
        resp.raise_for_status()
        return resp.json()

    def get_child_work_items(self, parent_id: int) -> list[dict[str, Any]]:
        """Fetch child work items linked to a parent via Hierarchy-Forward relations."""
        parent = self.get_work_item(parent_id, expand="relations")
        relations = parent.get("relations", []) or []

        child_urls = [
            rel["url"]
            for rel in relations
            if rel.get("rel") == "System.LinkTypes.Hierarchy-Forward"
        ]
        if not child_urls:
            return []

        children: list[dict[str, Any]] = []
        for url in child_urls:
            try:
                child_id = int(url.rstrip("/").split("/")[-1])
                child = self.get_work_item(child_id)
                children.append(child)
            except Exception as e:
                logger.warning("Failed to fetch child from %s: %s", url, e)
        return children

    def run_wiql(self, project: str, wiql: str) -> list[int]:
        url = f"{self._project_url(project)}/wit/wiql"
        params = {"api-version": ADO_API_VERSION}
        resp = self._client.post(
            url, params=params, json={"query": wiql},
            headers={"Content-Type": "application/json"},
        )
        resp.raise_for_status()
        rows = resp.json().get("workItems", [])
        return [row["id"] for row in rows]

    def run_saved_query(self, project: str, query_id: str) -> list[int]:
        url = f"{self._project_url(project)}/wit/wiql/{query_id}"
        params = {"api-version": ADO_API_VERSION}
        resp = self._client.get(url, params=params)
        resp.raise_for_status()
        data = resp.json()
        query_name = data.get("queryName", data.get("name", query_id))
        logger.info("Query name: %s", query_name)
        rows = data.get("workItems", [])
        return [row["id"] for row in rows]

    def create_work_item(
        self,
        project: str,
        work_item_type: str,
        fields: dict[str, Any],
        parent_id: int | None = None,
    ) -> dict[str, Any]:
        """Create a work item using JSON Patch format."""
        url = f"{self._project_url(project)}/wit/workitems/${work_item_type}"
        params = {"api-version": ADO_API_VERSION}

        operations: list[dict[str, Any]] = []
        for field_ref, value in fields.items():
            operations.append({
                "op": "add",
                "path": f"/fields/{field_ref}",
                "value": value,
            })

        if parent_id is not None:
            parent_url = f"https://dev.azure.com/{self._org}/_apis/wit/workitems/{parent_id}"
            operations.append({
                "op": "add",
                "path": "/relations/-",
                "value": {
                    "rel": "System.LinkTypes.Hierarchy-Reverse",
                    "url": parent_url,
                },
            })

        resp = self._client.post(url, params=params, json=operations)
        resp.raise_for_status()
        return resp.json()

    def create_child_work_item(
        self,
        project: str,
        parent_id: int,
        child_type: str,
        client_name: str,
        parent_fields: dict[str, Any],
    ) -> dict[str, Any]:
        """
        Create a child work item — same logic as ado_service.create_child_work_item.

        Sets title, client fields, and copies content + metadata fields from parent.
        """
        parent_title = parent_fields.get(TITLE, "Untitled")
        child_title = f"{client_name} - {parent_title}"

        child_fields: dict[str, Any] = {
            TITLE: child_title,
            CLIENT_REQUESTED: client_name,
            CLIENT_SELECTION_PORTAL: client_name,
        }

        # Determine which fields to sync based on parent type
        parent_type = parent_fields.get(WORK_ITEM_TYPE, "")
        content_fields, metadata_fields = _SYNC_FIELDS.get(parent_type, _DEFAULT_SYNC)

        # Copy content fields (no [TAG] filtering for bulk — full content)
        for field_ref in content_fields:
            value = parent_fields.get(field_ref)
            if value is not None:
                child_fields[field_ref] = value

        # Copy metadata fields as-is
        for field_ref in metadata_fields:
            value = parent_fields.get(field_ref)
            if value is not None:
                child_fields[field_ref] = value

        return self.create_work_item(
            project=project,
            work_item_type=child_type,
            fields=child_fields,
            parent_id=parent_id,
        )

    def update_work_item_state(self, work_item_id: int, state: str) -> None:
        """
        Set the state on a work item via a separate PATCH.

        State transitions must be done after creation because ADO enforces
        workflow rules that prevent setting arbitrary states at create time.
        """
        url = f"{self._base_url}/wit/workitems/{work_item_id}"
        params = {"api-version": ADO_API_VERSION}
        operations = [
            {"op": "add", "path": f"/fields/{STATE}", "value": state},
        ]
        resp = self._client.patch(url, params=params, json=operations)
        resp.raise_for_status()

    def get_attachments(self, work_item_id: int) -> list[dict[str, Any]]:
        """Get attachment metadata from a work item's relations."""
        wi = self.get_work_item(work_item_id, expand="relations")
        relations = wi.get("relations", []) or []
        return [
            {
                "url": rel["url"],
                "name": rel.get("attributes", {}).get("name", "unknown"),
            }
            for rel in relations
            if rel.get("rel") == "AttachedFile"
        ]

    def link_attachments_to_work_item(
        self, source_work_item_id: int, target_work_item_id: int
    ) -> int:
        """
        Link all attachments from one work item to another (no re-upload).

        ADO attachments live in a shared store — we just add AttachedFile
        relations on the target pointing to the same URLs.

        Returns the number of attachments successfully linked.
        """
        attachments = self.get_attachments(source_work_item_id)
        if not attachments:
            return 0

        linked = 0
        for att in attachments:
            filename = att.get("name", "unknown")
            att_url = att.get("url", "")
            try:
                url = f"{self._base_url}/wit/workitems/{target_work_item_id}"
                params = {"api-version": ADO_API_VERSION}
                operations = [
                    {
                        "op": "add",
                        "path": "/relations/-",
                        "value": {
                            "rel": "AttachedFile",
                            "url": att_url,
                            "attributes": {"name": filename},
                        },
                    }
                ]
                resp = self._client.patch(url, params=params, json=operations)
                resp.raise_for_status()
                linked += 1
            except Exception as e:
                logger.warning(
                    "  Failed to link attachment '%s' — skipping: %s",
                    filename, e,
                )
        return linked


# ---------------------------------------------------------------------------
# Existing children lookup
# ---------------------------------------------------------------------------

def _get_existing_children_by_client(
    ado: AdoClient,
    parent_id: int,
    expected_child_type: str,
) -> dict[str, int]:
    """Return {client_name_lower: child_id} for existing children of the expected type."""
    children = ado.get_child_work_items(parent_id)
    client_to_child: dict[str, int] = {}

    for child in children:
        child_fields = child.get("fields", {})
        child_type = child_fields.get(WORK_ITEM_TYPE, "")
        if child_type != expected_child_type:
            continue

        child_id = child["id"]
        client_name = child_fields.get(CLIENT_REQUESTED, "")
        if client_name:
            client_to_child[client_name.lower()] = child_id

    return client_to_child


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

def _generate_report(
    results: list[dict[str, Any]],
    output_path: str,
    org: str,
) -> None:
    """Generate an Excel spreadsheet from the creation results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Child Creation Report"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Parent ID",
        "Parent URL",
        "Parent Title",
        "Client",
        "Child Work Item ID",
        "Child Work Item URL",
        "Child Type",
        "Status",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    link_font = Font(name="Arial", color="0563C1", underline="single", size=10)
    data_font = Font(name="Arial", size=10)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, result in enumerate(results, 2):
        parent_id = result["parent_id"]
        project = result.get("project", "")
        parent_url = f"https://dev.azure.com/{org}/{project}/_workitems/edit/{parent_id}"
        child_id = result.get("child_id")
        child_url = (
            f"https://dev.azure.com/{org}/{project}/_workitems/edit/{child_id}"
            if child_id else ""
        )

        row_data = [
            parent_id,
            parent_url,
            result.get("parent_title", ""),
            result.get("client", ""),
            child_id or "",
            child_url,
            result.get("child_type", ""),
            result.get("status", ""),
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = wrap_alignment

        if parent_url:
            url_cell = ws.cell(row=row_idx, column=2)
            url_cell.hyperlink = parent_url
            url_cell.font = link_font
        if child_url:
            url_cell = ws.cell(row=row_idx, column=6)
            url_cell.hyperlink = child_url
            url_cell.font = link_font

    col_widths = [12, 55, 40, 25, 18, 55, 28, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(results) + 1}"

    wb.save(output_path)
    logger.info("Report saved to %s (%d rows)", output_path, len(results))


# ---------------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------------

def bulk_create(
    project: str,
    org: str,
    work_item_types: list[str] | None = None,
    states: list[str] | None = None,
    query_id: str = "",
    output_path: str = "bulk_children_report.xlsx",
    dry_run: bool = False,
) -> None:
    """Query ADO for parents and create child work items."""
    ado = AdoClient(org)
    results: list[dict[str, Any]] = []

    try:
        # 1. Query for matching parents
        if query_id:
            logger.info("Running saved query: %s", query_id)
            parent_ids = ado.run_saved_query(project, query_id)
        else:
            type_clauses = " OR ".join(
                f"[System.WorkItemType] = '{t}'" for t in (work_item_types or [])
            )
            state_clauses = " OR ".join(
                f"[System.State] = '{s}'" for s in (states or [])
            )
            wiql = (
                f"SELECT [System.Id] FROM workitems "
                f"WHERE [System.TeamProject] = '{project}' "
                f"AND ({type_clauses}) "
                f"AND ({state_clauses}) "
                f"ORDER BY [System.Id] ASC"
            )
            logger.info("Running WIQL query:\n  %s", wiql)
            parent_ids = ado.run_wiql(project, wiql)

        logger.info("Query returned %d work items", len(parent_ids))

        if not parent_ids:
            logger.info("No matching parent work items found. Nothing to do.")
            return

        logger.info(
            "Processing %d parent work items (dry_run=%s)",
            len(parent_ids), dry_run,
        )

        created_count = 0
        skipped_count = 0
        error_count = 0

        for i, parent_id in enumerate(parent_ids, 1):
            logger.info("--- [%d/%d] Parent %d ---", i, len(parent_ids), parent_id)

            wi_project = project  # default; overridden below if available
            fields: dict[str, Any] = {}
            client_name = ""

            try:
                # 2. Fetch the parent work item
                work_item = ado.get_work_item(parent_id)
                fields = work_item.get("fields", {})
                parent_type = fields.get(WORK_ITEM_TYPE, "")
                parent_title = fields.get(TITLE, "Untitled")
                wi_project = fields.get(TEAM_PROJECT, project)

                # 3. Determine the child type
                child_type = PARENT_TO_CHILD.get(parent_type)
                if not child_type:
                    logger.warning(
                        "  No child type mapping for parent type '%s' — skipping",
                        parent_type,
                    )
                    results.append({
                        "parent_id": parent_id,
                        "project": wi_project,
                        "parent_title": parent_title,
                        "client": "",
                        "child_id": None,
                        "child_type": "",
                        "status": f"SKIPPED — no child type for {parent_type}",
                    })
                    skipped_count += 1
                    continue

                # 4. Get clients from Client Selection for Portal Syncing
                portal_clients_raw = fields.get(CLIENT_SELECTION_PORTAL, "")
                if not portal_clients_raw or not str(portal_clients_raw).strip():
                    # Fall back to Client Requested if portal field is empty
                    portal_clients_raw = fields.get(CLIENT_REQUESTED, "")

                if not portal_clients_raw or not str(portal_clients_raw).strip():
                    logger.warning(
                        "  Parent %d has no portal clients or Client Requested — skipping",
                        parent_id,
                    )
                    results.append({
                        "parent_id": parent_id,
                        "project": wi_project,
                        "parent_title": parent_title,
                        "client": "",
                        "child_id": None,
                        "child_type": child_type,
                        "status": "SKIPPED — no clients",
                    })
                    skipped_count += 1
                    continue

                # Portal field can be semicolon-delimited for multiple clients
                client_names = [
                    c.strip() for c in str(portal_clients_raw).split(";")
                    if c.strip()
                ]

                # 5. Check for existing children
                existing = _get_existing_children_by_client(ado, parent_id, child_type)

                for client_name in client_names:
                    if client_name.lower() in existing:
                        existing_id = existing[client_name.lower()]
                        logger.info(
                            "  Child already exists for client '%s' (ID %d) — skipping",
                            client_name, existing_id,
                        )
                        results.append({
                            "parent_id": parent_id,
                            "project": wi_project,
                            "parent_title": parent_title,
                            "client": client_name,
                            "child_id": existing_id,
                            "child_type": child_type,
                            "status": "SKIPPED — child already exists",
                        })
                        skipped_count += 1
                        continue

                    # 6. Create the child work item
                    if dry_run:
                        logger.info(
                            "  [DRY RUN] Would create %s for client '%s'",
                            child_type, client_name,
                        )
                        results.append({
                            "parent_id": parent_id,
                            "project": wi_project,
                            "parent_title": parent_title,
                            "client": client_name,
                            "child_id": None,
                            "child_type": child_type,
                            "status": "DRY RUN — would create",
                        })
                        continue

                    new_child = ado.create_child_work_item(
                        project=wi_project,
                        parent_id=parent_id,
                        child_type=child_type,
                        client_name=client_name,
                        parent_fields=fields,
                    )
                    new_child_id = new_child.get("id")

                    logger.info(
                        "  Created %s (ID %d) for client '%s'",
                        child_type, new_child_id, client_name,
                    )

                    # 6a. Sync parent state → child state
                    parent_state = fields.get(STATE, "")
                    if parent_state:
                        state_map = STATE_TO_CHILD_STATE.get(parent_type, {})
                        # For types with a mapping (Epic, Initiative, Feature),
                        # look up the child state. For 1:1 types (User Story,
                        # Bug), the map is empty — copy parent state directly.
                        child_state = state_map.get(parent_state, parent_state) if state_map else parent_state
                        try:
                            ado.update_work_item_state(new_child_id, child_state)
                            logger.info(
                                "  Set child state to '%s' (parent was '%s')",
                                child_state, parent_state,
                            )
                        except Exception as e:
                            logger.warning(
                                "  Failed to set child state to '%s' — skipping: %s",
                                child_state, e,
                            )

                    # 6b. Link parent attachments to child (shared store, no re-upload)
                    try:
                        att_count = ado.link_attachments_to_work_item(
                            source_work_item_id=parent_id,
                            target_work_item_id=new_child_id,
                        )
                        if att_count:
                            logger.info(
                                "  Linked %d attachment(s) from parent to child",
                                att_count,
                            )
                    except Exception as e:
                        logger.warning(
                            "  Failed to link attachments — skipping: %s", e,
                        )

                    results.append({
                        "parent_id": parent_id,
                        "project": wi_project,
                        "parent_title": parent_title,
                        "client": client_name,
                        "child_id": new_child_id,
                        "child_type": child_type,
                        "status": "CREATED",
                    })
                    created_count += 1

            except Exception as e:
                logger.error("  Error processing parent %d: %s", parent_id, e)
                results.append({
                    "parent_id": parent_id,
                    "project": wi_project,
                    "parent_title": fields.get(TITLE, ""),
                    "client": client_name,
                    "child_id": None,
                    "child_type": "",
                    "status": f"ERROR — {str(e)[:200]}",
                })
                error_count += 1

        # 7. Summary
        logger.info("=" * 60)
        logger.info("SUMMARY")
        logger.info("  Total parents queried:  %d", len(parent_ids))
        logger.info("  Children created:       %d", created_count)
        logger.info("  Skipped (existing/no client): %d", skipped_count)
        logger.info("  Errors:                 %d", error_count)
        logger.info("=" * 60)

        # 8. Generate report
        if results:
            _generate_report(results, output_path, org)

    finally:
        ado.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Bulk create client child work items from parent ADO work items.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--project",
        required=True,
        help="ADO project name (e.g., 'MyProject')",
    )
    parser.add_argument(
        "--types",
        help="Comma-separated parent work item types (e.g., 'User Story,Issue,Bug,Epic'). "
             "Required unless --query-id is provided.",
    )
    parser.add_argument(
        "--states",
        help="Comma-separated work item states to match (e.g., 'Ready for Development,In Progress'). "
             "Required unless --query-id is provided.",
    )
    parser.add_argument(
        "--query-id",
        help="GUID of a saved ADO query (flat list type). When provided, --types and "
             "--states are ignored and all work items returned by the query are processed.",
    )
    parser.add_argument(
        "--output",
        default="bulk_children_report.xlsx",
        help="Output Excel file path (default: bulk_children_report.xlsx)",
    )
    parser.add_argument(
        "--org",
        default=DEFAULT_ORG,
        help=f"ADO organization name (default: {DEFAULT_ORG})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview what would be created without making changes",
    )

    args = parser.parse_args()

    # Validate: either --query-id OR (--types AND --states) must be provided
    if not args.query_id:
        if not args.types:
            parser.error("--types is required when --query-id is not provided")
        if not args.states:
            parser.error("--states is required when --query-id is not provided")

    work_item_types: list[str] = []
    states: list[str] = []

    if args.types:
        work_item_types = [t.strip() for t in args.types.split(",") if t.strip()]
    if args.states:
        states = [s.strip() for s in args.states.split(",") if s.strip()]

    if not args.query_id:
        if not work_item_types:
            parser.error("--types must include at least one work item type")
        if not states:
            parser.error("--states must include at least one state")

    # Validate types are known parent types
    for t in work_item_types:
        if t not in ALL_PARENT_TYPES:
            logger.warning(
                "Warning: '%s' is not a recognized parent type. Known types: %s",
                t, ", ".join(sorted(ALL_PARENT_TYPES)),
            )

    logger.info("Bulk Child Creation Script")
    logger.info("  Organization: %s", args.org)
    logger.info("  Project:      %s", args.project)
    if args.query_id:
        logger.info("  Query ID:     %s", args.query_id)
    else:
        logger.info("  Types:        %s", work_item_types)
        logger.info("  States:       %s", states)
    logger.info("  Output:       %s", args.output)
    logger.info("  Dry run:      %s", args.dry_run)
    logger.info("")

    bulk_create(
        project=args.project,
        org=args.org,
        work_item_types=work_item_types or None,
        states=states or None,
        query_id=args.query_id or "",
        output_path=args.output,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()

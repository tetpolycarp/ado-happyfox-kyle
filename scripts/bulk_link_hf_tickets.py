"""
Bulk Link HappyFox Tickets to ADO Client Work Items.

Self-contained one-off script that reads a CSV mapping ADO Work Item IDs to
HappyFox Ticket IDs, fetches each HF ticket's subject, and updates the
corresponding ADO work item's Support Ticket Number and Support Ticket Title
fields.

Input CSV format (first row is header):
    Work Item URL,Work Item ID,Happy Fox Ticket ID,Response Code,Response Message

For each row the script:
    1. Fetches the HappyFox ticket (GET /api/1.1/json/ticket/{id}/)
    2. Extracts the ticket subject
    3. Patches the ADO work item with:
       - Custom.SupportTicketNumber = HF ticket ID (string)
       - Custom.SupportTicketTitle  = HF ticket subject

Outputs an Excel report with results for each row.

Usage
-----
  # Dry run — preview what would be updated
  python scripts/bulk_link_hf_tickets.py \\
    --csv "path/to/mapping.csv" \\
    --dry-run

  # Run for real
  python scripts/bulk_link_hf_tickets.py \\
    --csv "path/to/mapping.csv"

  # Override HF URL or ADO org
  python scripts/bulk_link_hf_tickets.py \\
    --csv "path/to/mapping.csv" \\
    --hf-url "https://brandtinfo.happyfox.com/api/1.1/json" \\
    --org BrandtInfoServices

Authentication
--------------
- ADO: DefaultAzureCredential (Azure CLI locally, Managed Identity in Azure).
  Requires: `az login` before running locally.
- HappyFox: HTTP Basic Auth via environment variables:
    HF_API_KEY    — HappyFox API key
    HF_AUTH_CODE  — HappyFox auth code

Dependencies
------------
  pip install httpx openpyxl azure-identity
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import time
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("bulk_link_hf_tickets")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ADO_API_VERSION = "7.1"
ADO_RESOURCE_ID = "499b84ac-1321-427f-aa17-267ca6975798"
TOKEN_REFRESH_BUFFER_SECONDS = 300
DEFAULT_ORG = os.environ.get("ADO_ORGANIZATION", "BrandtInfoServices")
DEFAULT_HF_URL = os.environ.get(
    "HF_API_URL", "https://brandtinfo.happyfox.com/api/1.1/json"
)

# ADO field reference names
SUPPORT_TICKET_NUMBER = os.environ.get(
    "ADO_FIELD_SUPPORT_TICKET_NUMBER", "Custom.SupportTicketNumber"
)
SUPPORT_TICKET_TITLE = os.environ.get(
    "ADO_FIELD_SUPPORT_TICKET_TITLE", "Custom.SupportTicketTitle"
)


# ---------------------------------------------------------------------------
# ADO HTTP client (same pattern as other bulk scripts)
# ---------------------------------------------------------------------------

class AdoClient:
    """Minimal ADO REST API client using DefaultAzureCredential."""

    def __init__(self, org: str) -> None:
        self._org = org
        self._base_url = f"https://dev.azure.com/{org}/_apis"
        self._credential = None
        self._cached_token: str | None = None
        self._token_expires_at: float = 0.0
        self._client = httpx.Client(
            headers={"Content-Type": "application/json-patch+json"},
            timeout=30.0,
        )

    def _ensure_token(self) -> str:
        now = time.time()
        if self._cached_token and now < self._token_expires_at:
            return self._cached_token

        if self._credential is None:
            from azure.identity import DefaultAzureCredential
            self._credential = DefaultAzureCredential()

        token = self._credential.get_token(f"{ADO_RESOURCE_ID}/.default")
        self._cached_token = token.token
        self._token_expires_at = token.expires_on - TOKEN_REFRESH_BUFFER_SECONDS
        logger.info("ADO token refreshed (expires in %ds)", int(token.expires_on - now))
        return self._cached_token

    def _auth_headers(self) -> dict[str, str]:
        return {"Authorization": f"Bearer {self._ensure_token()}"}

    def update_work_item(
        self, work_item_id: int, fields: dict[str, Any], project: str = ""
    ) -> dict:
        """PATCH a work item with the given field values."""
        # Build JSON Patch document
        patch_doc = [
            {"op": "add", "path": f"/fields/{ref}", "value": value}
            for ref, value in fields.items()
            if value is not None
        ]
        if not patch_doc:
            return {}

        if project:
            url = f"https://dev.azure.com/{self._org}/{project}/_apis/wit/workitems/{work_item_id}"
        else:
            url = f"{self._base_url}/wit/workitems/{work_item_id}"

        resp = self._client.patch(
            url,
            json=patch_doc,
            headers=self._auth_headers(),
            params={"api-version": ADO_API_VERSION},
        )
        resp.raise_for_status()
        return resp.json()

    def close(self) -> None:
        self._client.close()


# ---------------------------------------------------------------------------
# HappyFox HTTP client
# ---------------------------------------------------------------------------

class HfClient:
    """Minimal HappyFox REST API client using Basic Auth.

    Includes rate-limit handling: HappyFox allows 500 GET requests per
    minute. The client tracks calls in a sliding window and sleeps when
    approaching the limit. If a 429 is received, it respects Retry-After.
    """

    # Stay safely under the 500/min ceiling
    _MAX_CALLS_PER_WINDOW = 450
    _WINDOW_SECONDS = 60.0

    def __init__(self, base_url: str, api_key: str, auth_code: str) -> None:
        self._base_url = base_url.rstrip("/")
        self._client = httpx.Client(
            auth=(api_key, auth_code),
            timeout=30.0,
        )
        self._call_times: list[float] = []

    def _throttle(self) -> None:
        """Sleep if we're approaching the rate limit."""
        now = time.time()
        cutoff = now - self._WINDOW_SECONDS
        # Prune calls outside the sliding window
        self._call_times = [t for t in self._call_times if t > cutoff]

        if len(self._call_times) >= self._MAX_CALLS_PER_WINDOW:
            # Wait until the oldest call in the window expires
            sleep_for = self._call_times[0] - cutoff + 1.0
            logger.info("Rate-limit throttle: sleeping %.1fs", sleep_for)
            time.sleep(sleep_for)
            # Prune again after sleeping
            now = time.time()
            cutoff = now - self._WINDOW_SECONDS
            self._call_times = [t for t in self._call_times if t > cutoff]

    def get_ticket(self, ticket_id: int, max_retries: int = 3) -> dict:
        """Fetch a single HappyFox ticket by ID with rate-limit retry."""
        self._throttle()

        for attempt in range(1, max_retries + 1):
            self._call_times.append(time.time())
            resp = self._client.get(f"{self._base_url}/ticket/{ticket_id}/")

            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", "60"))
                logger.warning(
                    "HF 429 rate-limited (attempt %d/%d) — waiting %ds",
                    attempt, max_retries, retry_after,
                )
                time.sleep(retry_after)
                continue

            resp.raise_for_status()
            return resp.json()

        # All retries exhausted
        raise httpx.HTTPStatusError(
            f"HF rate-limited after {max_retries} retries",
            request=resp.request,
            response=resp,
        )

    def close(self) -> None:
        self._client.close()


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

REPORT_COLUMNS = [
    "ADO Work Item ID",
    "HF Ticket ID",
    "HF Subject",
    "Support Ticket Number Set",
    "Support Ticket Title Set",
    "Status",
    "Error",
]


def _create_report(results: list[dict], output_path: str) -> None:
    """Write results to an Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Link Results"

    # Header row
    for col_idx, header in enumerate(REPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=r.get("ado_id", ""))
        ws.cell(row=row_idx, column=2, value=r.get("hf_id", ""))
        ws.cell(row=row_idx, column=3, value=r.get("hf_subject", ""))
        ws.cell(row=row_idx, column=4, value=r.get("ticket_number_set", ""))
        ws.cell(row=row_idx, column=5, value=r.get("ticket_title_set", ""))
        ws.cell(row=row_idx, column=6, value=r.get("status", ""))
        ws.cell(row=row_idx, column=7, value=r.get("error", ""))

        # Color-code the row based on status
        status = r.get("status", "")
        fill = SUCCESS_FILL if status == "OK" else ERROR_FILL if status == "ERROR" else SKIP_FILL
        for col_idx in range(1, len(REPORT_COLUMNS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    # Auto-width columns
    for col_idx in range(1, len(REPORT_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(results) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info("Report saved → %s (%d rows)", output_path, len(results))


# ---------------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------------

def _load_completed_ids(report_path: str) -> set[str]:
    """Load ADO work item IDs already processed (OK or DRY_RUN) from a prior report."""
    completed: set[str] = set()
    if not os.path.exists(report_path):
        return completed

    from openpyxl import load_workbook
    wb = load_workbook(report_path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: ADO Work Item ID, HF Ticket ID, HF Subject,
        #          Support Ticket Number Set, Support Ticket Title Set, Status, Error
        ado_id = str(row[0] or "").strip()
        status = str(row[5] or "").strip()
        if ado_id and status in ("OK", "DRY_RUN"):
            completed.add(ado_id)
    wb.close()
    logger.info("Resuming — %d work items already processed in %s", len(completed), report_path)
    return completed


def bulk_link(
    csv_path: str,
    org: str,
    hf_url: str,
    output_path: str,
    dry_run: bool,
    resume: bool,
) -> None:
    """Read the CSV, fetch HF ticket subjects, update ADO work items."""

    # Read HF credentials from environment
    hf_api_key = os.environ.get("HF_API_KEY", "")
    hf_auth_code = os.environ.get("HF_AUTH_CODE", "")
    if not hf_api_key or not hf_auth_code:
        raise SystemExit(
            "Missing HF credentials. Set HF_API_KEY and HF_AUTH_CODE environment variables."
        )

    # Load already-completed IDs if resuming
    already_done: set[str] = set()
    if resume:
        already_done = _load_completed_ids(output_path)

    # Parse CSV
    rows: list[dict[str, str]] = []
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    logger.info("Loaded %d rows from %s", len(rows), csv_path)
    if already_done:
        logger.info("Will skip %d already-completed rows", len(already_done))

    if not rows:
        logger.warning("No rows to process — exiting")
        return

    ado = AdoClient(org)
    hf = HfClient(hf_url, hf_api_key, hf_auth_code)
    results: list[dict] = []

    try:
        for i, row in enumerate(rows, start=1):
            ado_id_str = row.get("Work Item ID", "").strip()
            hf_id_str = row.get("Happy Fox Ticket ID", "").strip()

            if not ado_id_str or not hf_id_str:
                logger.warning("Row %d: missing ADO ID or HF ID — skipping", i)
                results.append({
                    "ado_id": ado_id_str,
                    "hf_id": hf_id_str,
                    "hf_subject": "",
                    "ticket_number_set": "",
                    "ticket_title_set": "",
                    "status": "SKIP",
                    "error": "Missing ADO Work Item ID or HF Ticket ID",
                })
                continue

            if ado_id_str in already_done:
                continue

            ado_id = int(ado_id_str)
            hf_id = int(hf_id_str)

            # Step 1: Fetch HF ticket to get subject
            try:
                hf_ticket = hf.get_ticket(hf_id)
                hf_subject = hf_ticket.get("subject", "")
            except Exception as e:
                logger.error("Row %d: failed to fetch HF ticket %d — %s", i, hf_id, e)
                results.append({
                    "ado_id": ado_id,
                    "hf_id": hf_id,
                    "hf_subject": "",
                    "ticket_number_set": "",
                    "ticket_title_set": "",
                    "status": "ERROR",
                    "error": f"HF fetch failed: {str(e)[:200]}",
                })
                continue

            logger.info(
                "Row %d/%d: ADO %d ← HF #%d \"%s\"",
                i, len(rows), ado_id, hf_id, hf_subject[:60],
            )

            if dry_run:
                results.append({
                    "ado_id": ado_id,
                    "hf_id": hf_id,
                    "hf_subject": hf_subject,
                    "ticket_number_set": str(hf_id),
                    "ticket_title_set": hf_subject,
                    "status": "DRY_RUN",
                    "error": "",
                })
                continue

            # Step 2: Update ADO work item
            try:
                ado.update_work_item(ado_id, {
                    SUPPORT_TICKET_NUMBER: str(hf_id),
                    SUPPORT_TICKET_TITLE: hf_subject,
                })
                results.append({
                    "ado_id": ado_id,
                    "hf_id": hf_id,
                    "hf_subject": hf_subject,
                    "ticket_number_set": str(hf_id),
                    "ticket_title_set": hf_subject,
                    "status": "OK",
                    "error": "",
                })
            except Exception as e:
                logger.error("Row %d: failed to update ADO %d — %s", i, ado_id, e)
                results.append({
                    "ado_id": ado_id,
                    "hf_id": hf_id,
                    "hf_subject": hf_subject,
                    "ticket_number_set": "",
                    "ticket_title_set": "",
                    "status": "ERROR",
                    "error": f"ADO update failed: {str(e)[:200]}",
                })

    finally:
        ado.close()
        hf.close()

    # Summary
    ok = sum(1 for r in results if r["status"] == "OK")
    dry = sum(1 for r in results if r["status"] == "DRY_RUN")
    err = sum(1 for r in results if r["status"] == "ERROR")
    skip = sum(1 for r in results if r["status"] == "SKIP")
    logger.info(
        "Done: %d OK, %d dry-run, %d errors, %d skipped (of %d total)",
        ok, dry, err, skip, len(results),
    )

    _create_report(results, output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Bulk link HappyFox tickets to ADO work items by setting "
                    "Support Ticket Number and Support Ticket Title.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--csv",
        required=True,
        help="Path to CSV file with columns: Work Item ID, Happy Fox Ticket ID",
    )
    parser.add_argument(
        "--org",
        default=DEFAULT_ORG,
        help=f"ADO organization name (default: {DEFAULT_ORG})",
    )
    parser.add_argument(
        "--hf-url",
        default=DEFAULT_HF_URL,
        help=f"HappyFox API base URL (default: {DEFAULT_HF_URL})",
    )
    parser.add_argument(
        "--output",
        default="bulk_link_hf_report.xlsx",
        help="Output Excel report path (default: bulk_link_hf_report.xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Fetch HF tickets but don't update ADO — preview only",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Skip rows already processed (OK/DRY_RUN) in a previous output report. "
             "Reads the --output file to determine which ADO IDs are done.",
    )

    args = parser.parse_args()
    bulk_link(
        csv_path=args.csv,
        org=args.org,
        hf_url=args.hf_url,
        output_path=args.output,
        dry_run=args.dry_run,
        resume=args.resume,
    )


if __name__ == "__main__":
    main()
